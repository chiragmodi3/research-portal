import re
import pandas as pd
from pdf2image import convert_from_path
import pytesseract
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

def detect_year_columns(ocr_text):
    matches = re.findall(r"(20\d{2})", ocr_text)

    years = []
    seen = set()

    for y in matches:
        if y not in seen:
            seen.add(y)
            years.append("FY" + y[-2:])

    return years[:6]

def extract_financial_tables_fast(pdf_path):

    images = convert_from_path(
        pdf_path,
        dpi=150,
        first_page=1,
        last_page=1,
        grayscale=True
    )

    img = images[0]

    data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DATAFRAME)

    data = data[data.conf != -1]
    data = data.dropna(subset=["text"])

    rows = []

    current_line = []
    last_top = None

    for _, row in data.iterrows():

        if last_top is None:
            last_top = row["top"]

        if abs(row["top"] - last_top) > 10:
            if current_line:
                rows.append(current_line)
            current_line = []
            last_top = row["top"]

        current_line.append((row["left"], row["text"]))

    if current_line:
        rows.append(current_line)

    table_rows = []

    for line in rows:
        line_sorted = sorted(line, key=lambda x: x[0])
        text_line = " ".join([w[1] for w in line_sorted])

        numbers = re.findall(r"\(?-?\d[\d,]*\.?\d*\)?", text_line)

        label = re.sub(r"\(?-?\d[\d,]*\.?\d*\)?", "", text_line).strip()

        if numbers:
            table_rows.append([label] + numbers)
        else:
            table_rows.append([label])

    df = pd.DataFrame(table_rows)

    financial_keywords = [
        "revenue",
        "income",
        "expense",
        "profit",
        "tax",
        "cost",
        "ebitda",
        "loss"
    ]

    df = df[df[0].astype(str).str.lower().apply(
        lambda x: any(k in x for k in financial_keywords)
    )]

    return df, ["FY25", "FY24"]

def format_financial_excel(file_path, detected_years=None):

    wb = load_workbook(file_path)
    ws = wb.active

    headers = ["Particulars"]

    if detected_years:
        headers.extend(detected_years)
    else:
        headers.extend(["FY25", "FY24", "FY23", "FY22", "FY21", "FY20"])

    for col in range(1, ws.max_column + 1):
        if col <= len(headers):
            ws.cell(row=1, column=col).value = headers[col - 1]

    bold = Font(bold=True)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    section_keywords = [
        "income", "revenue", "expenses", "profit",
        "loss", "ebitda", "tax"
    ]

    for row in range(2, ws.max_row + 1):

        label_cell = ws.cell(row=row, column=1)
        label = str(label_cell.value).lower() if label_cell.value else ""

        if any(keyword in label for keyword in section_keywords):
            label_cell.font = bold

        for col in range(2, ws.max_column + 1):

            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="right")

            try:
                if cell.value:
                    val = str(cell.value).replace(",", "")
                    float(val)
                    cell.number_format = '#,##0'
            except:
                pass

    ws.column_dimensions["A"].width = 45

    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.freeze_panes = "A2"

    wb.save(file_path)
